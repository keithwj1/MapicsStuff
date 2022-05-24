from jira import JIRA, JIRAError
import sys
import json
#import ibm_db
import pyodbc
import logging
import logging.handlers
import os
from openpyxl import Workbook, load_workbook
import getpass
import datetime
#import os.path
#from pathlib import Path
from os import path
sys.path.append(os.path.expanduser('~\source\repos\Mapics'))
from MapicsClasses import *




class cMapicsData:
    def __init__(self):
        self.PurchaseOrderList = cPurchaseOrderList();
        self.ManufacturingOrderList = cManufacturingOrderList();
        self.RouterOppList = cRouterList();
        self.SupplierList = cSupplierList();
        self.ItemLocations = cItemLocationList();
        self.PartList = cPartList();
        self.BuyerList = cBuyerList();
        self.bFilled = False;
    def FillMapicsData(self):
        if self.bFilled == False:
            self.TestMapicsConnection();
    def TestMapicsConnection(self):
        server = 'tcp:ATLASSIAN' 
        database = 'master' 
        username = 'jiramapics' 
        password = '' 
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cursor = cnxn.cursor()
        sqlquery = "SELECT TOP 1 [BUYNO] FROM [MAPICS].[POWER9].[AMFLIBF].[BUYERF]"
        try:
            cursor.execute(sqlquery) 
        except cursor.Error as e:
            cursor.close()
            cnxn.close()
            sqlautoitprogram = os.path.expanduser('~\Documents\AutoIt Script\sqlpropertiesopen.exe')
            os.system(sqlautoitprogram)
            logging.exception("Running SQL Login Program")

            sys.exit(0)


class cDefect:
    def __init__(self,nItem = 0, nQty = 0, sDwgZone = "",sDescription = "",sShouldBe = "SRMRA SHOULD BE",sActual = "SRMRA ACTUAL"):
        self.nItem = nItem
        self.nQty = nQty
        self.sDwgZone = sDwgZone
        self.sDescription = sDescription
        if (sShouldBe == ""):
            sShouldBe = "SRMRA SHOULD BE"
        self.sShouldBe = sShouldBe
        if (sActual == ""):
            sActual = "SRMRA ACTUAL"
        self.sActual = sActual
jira_return = ""
jira_user = ""
while (jira_user == "" and jira_user != "q"):
    jira_user = input("Enter UserName: ")
if (jira_user == "q"):
    sys.exit()
jira_password = ""
if( jira_user == ""):
    jira_user = "ImportUser"
else:
    jira_password = getpass.getpass('Password:')

jira_server = "https://atlassian/jira"
jira_server = {'server': jira_server}
try:
    jira = JIRA(options=jira_server, basic_auth=(jira_user, jira_password))
except JIRAError as e:
    if e.status_code == 401:
        print("Login to JIRA failed. Check your username and password")
        sys.exit()
    else:
        print("Unknown Error " + str(e.status_code) + " Response = "+ str(e.response))
        sys.exit()

if len(sys.argv) > 1:
    ifile = sys.argv[1]
else:
    ifile = input("Please Drop File on Window. Then press Enter \n")
    #ifile = input("Please Drop File on EXE. Press Enter To Exit \n")
if (path.exists(ifile) == False):
    ifile = ifile.replace('"','')
    if (path.exists(ifile) == False):
        input("File does not exist, Press enter to exit")
        sys.exit()
SRMRABook = load_workbook(ifile)
try:
    SRMRASheet = SRMRABook['Q1033']
except:
    print("Sheet with name Q1033 does not exist. Exiting")
    input("")
    sys.exit()

sCustomerApproval = input ("Is Customer Approval Required? (Yes/No) \n")
if "n" in sCustomerApproval.lower():
    sCustomerApproval = "No"
else:
    sCustomerApproval = "Yes"

sBussinessUnit = input ("Which Business Unit? (EM/ESPG/GCU) \n")
if "g" in sBussinessUnit.lower():
    sBussinessUnit = "GCU"
elif "s" in sBussinessUnit.lower():
    sBussinessUnit = "ESPG"
else:
    sBussinessUnit = "EM"




print ("Import SRMRA DATA")
data = {}
sPurchaseOrder= ""
sDate = datetime.datetime.utcnow()
sAddress = ""
sPartNumber = ""
sRevision = ""
sPartDescription = ""
sLineItem = ""
sTotalQuantity = ""
sTotalQuantityDefective = ""
sRootCause = ""
sCorrectiveAction = ""
sRequestor = ""
DefectList = []
nMaxDefects = 7
#determine max defects
nStartDefects = 0
nEndDefects = 0
for row in range(1,SRMRASheet.max_row):
    for col in range(1,SRMRASheet.max_column):
        ColName = SRMRASheet.cell(row,col).value
        if ColName == "QTY":
            nStartDefects = row
        elif ColName == "* MATERIAL DISPOSITIONED AS REPAIR SHALL BE REPAIRED IN ACCORDANCE WITH SKURKA AEROSPACE APPROVED REPAIR PROCEDURES.  VENDOR SHALL SUBMIT PROPOSED REPAIR PROCEDURE AT THE TIME OF RMRA SUBMITTAL FOR APPROVAL.":
            nEndDefects = row
nMaxDefects = nEndDefects - nStartDefects - 1
if nMaxDefects < 1:
    nMaxDefects = 3
for i in range(0,nMaxDefects):
    Defect = cDefect();
    DefectList.append(Defect)

for row in range(1,SRMRASheet.max_row):
    if row == SRMRASheet.min_row:
        continue;
    else:        
        row_below = row+1
        
        
        sValue = ""
        for col in range(1,SRMRASheet.max_column):
            cf = ""
            ColName = str(SRMRASheet.cell(row,col).value)
            if ColName == 'REQUESTED BY (VENDOR NAME & ADDRESS):':
                sValue = str(SRMRASheet.cell(row_below,col).value)
                sAddress = sValue
                cf = 'customfield_11818'
            elif ColName == 'SKURKA AEROSPACE PART NUMBER:':
                sValue = str(SRMRASheet.cell(row_below,col).value)
                sPartNumber = sValue
                cf = 'customfield_10705'
            elif ColName == 'REVISION:':
                sValue = str(SRMRASheet.cell(row_below,col).value)
                sRevision = sValue
                cf = 'customfield_11003'
            elif ColName == 'PART DESCRIPTION:':
                sValue = str(SRMRASheet.cell(row_below,col).value)
                sPartDescription = sValue             
                cf = 'customfield_11062'
            elif ColName == 'PURCHASE ORDER NO.:':
                sPurchaseOrder = str(SRMRASheet.cell(row_below,col).value)
                sPurchaseOrder = sPurchaseOrder.upper()
                sPurchaseOrder=sPurchaseOrder.replace(" ","")
                sValue = sPurchaseOrder
                cf = 'customfield_11215'
            elif ColName == 'LINE ITEM:':
                sValue = str(SRMRASheet.cell(row_below,col).value)
                sLineItem = sValue
                
                cf = "customfield_15201"
            elif ColName == 'QUANTITY DUE ON PURCHASE ORDER:':
                sValue = int(SRMRASheet.cell(row_below,col).value)
                sTotalQuantity = sValue
                cf = 'customfield_13232'
            elif ColName == 'QUANTITY DEFECTIVE:':
                sValue = int(SRMRASheet.cell(row_below,col).value)
                sTotalQuantityDefective = sValue
                cf = 'customfield_13231'
            elif ColName == 'ROOT CAUSE OF NONCONFORMANCE:':
                sValue = str(SRMRASheet.cell(row_below,col).value)
                sRootCause = sValue
                cf = 'customfield_12911'
            elif ColName == 'CORRECTIVE ACTION TAKEN TO PREVENT RECURRENCE:':
                sValue = str(SRMRASheet.cell(row_below,col).value)
                sCorrectiveAction = sValue 
                cf = 'customfield_12916'
            elif ColName == 'PERSON REQUESTING:':
                sValue = str(SRMRASheet.cell(row_below,col).value)
                sRequestor = sValue
                cf = 'customfield_11817'
            elif ColName == 'DATE:':
                sValue = SRMRASheet.cell(row_below,col).value
                sDate = sValue
                #cf = 'created'
            elif ColName == 'ITEM':
                for i in range(0,nMaxDefects):
                    if SRMRASheet.cell(row_below+i,col).value:
                        DefectList[i].nItem = SRMRASheet.cell(row_below+i,col).value
                    else:
                        DefectList[i].nItem = i+1
            elif ColName == 'QTY':
                for i in range(0,nMaxDefects):
                    if SRMRASheet.cell(row_below+i,col).value:
                        DefectList[i].nQty = int(SRMRASheet.cell(row_below+i,col).value)
                cf = 'customfield_13220'
            elif ColName == 'ZONE':
                for i in range(0,nMaxDefects):
                    if SRMRASheet.cell(row_below+i,col).value:
                        DefectList[i].sDwgZone = str(SRMRASheet.cell(row_below+i,col).value)
                cf = 'customfield_11029'
            elif ColName == 'DISCREPANCIES (DESCRIBE FULLY)':
                for i in range(0,nMaxDefects):
                    if SRMRASheet.cell(row_below+i,col).value:
                        DefectList[i].sDescription = str(SRMRASheet.cell(row_below+i,col).value)
                cf = 'customfield_11203'
            elif ColName == 'SHOULD BE':
                for i in range(0,nMaxDefects):
                    if SRMRASheet.cell(row_below+i,col).value:
                        DefectList[i].sShouldBe = str(SRMRASheet.cell(row_below+i,col).value)
                cf = 'customfield_11034'
            elif ColName == 'ACTUAL':
                for i in range(0,nMaxDefects):
                    if SRMRASheet.cell(row_below+i,col).value:
                        DefectList[i].sActual = str(SRMRASheet.cell(row_below+i,col).value)
                cf = 'customfield_11035'
            if sValue != "" and cf != "":
                data[cf] = sValue
key = SRMRASheet.cell(1,8).value
SRMRABook.close()

ThisPO = None
ThisSupplier = None
ThisBuyer = None
if len(sPurchaseOrder) > 5 and len(sPurchaseOrder) < 8:
    MapicsData = cMapicsData();
    #removed since Wil Ramos disabled the link to mapics
    #MapicsData.SupplierList.FillMapicsData()
    #aPO = MapicsData.PurchaseOrderList.FindAllPO(sPurchaseOrder,True,True,True,MapicsData)
    aPO = None
    if aPO != None:
        for PO in aPO:
            if (PO.sPartNumber.startswith("*") or PO.sPartNumber.startswith("OSP")):
                continue;
            else:
                ThisPO = PO
                Supplier = MapicsData.SupplierList.FindSupplier(PO.sSupplierNumber);
                data['customfield_10705'] = PO.sPartNumber
                data["summary"] = PO.sPartNumber
                data['customfield_11003'] = PO.sRevision
                data['customfield_11062'] = PO.sPartDescription
                data["customfield_15100"] = PO.sSupplierNumber
                data["customfield_15501"] = PO.sBuyerNumber
                data["customfield_14001"] = PO.sEngNumber
                if Supplier != None:
                    ThisSupplier = Supplier
                    data["customfield_11818"] = Supplier.sAddress + "\r\n" + Supplier.sPhone
                    data["customfield_11202"] = Supplier.sName
                    data["customfield_16100"] = Supplier.sEmail
                Buyer = MapicsData.BuyerList.FindBuyer(PO.sBuyerNumber)
                if Buyer != None:
                    ThisBuyer = Buyer
                    data["customfield_15503"] = Buyer.sEmail

try:
    issue = jira.issue(key)
except JIRAError as e:
    issue_dict = {
        'project': {'id': 12501},
        'summary': sPartNumber,
        'description': '',
        'issuetype': {'name': 'Task'},
    }
    issue = jira.create_issue(fields=issue_dict)


print("Updating Issue")  
data["customfield_11001"] = {'value':sBussinessUnit}
issue.update(fields=data)



issuestatus = str(issue.fields.status)
if issuestatus == "Open":
    #Create NCMR Transition
    print("Transition Create NCMR")  
    jira.transition_issue(issue, transition='11')

#Refresh issue details
issue = jira.issue(issue.key)

print("Get NCMR Issue")  

for link in issue.fields.issuelinks:
    linkkey = ""
    if hasattr(link, "outwardIssue"):
        outwardIssue = link.outwardIssue
        print("\tOutward: " + outwardIssue.key)
        linkkey = outwardIssue.key
    if hasattr(link, "inwardIssue"):
        inwardIssue = link.inwardIssue
        print("\tInward: " + inwardIssue.key)
        linkkey = inwardIssue.key
    if linkkey != "":
        linkissue = jira.issue(linkkey)
        if linkissue.fields.issuetype.name == "NCMR":
            #Get Defects
            DescriptionList = []
            if linkissue.fields.subtasks != None and 1==2:
                #Update defects????
                for subtasks in linkissue.fields.subtasks:
                    subtaskID = subtasks.key
                    subtaskissue = jira.issue(subtasks.key)
                    sDescription = getattr(subtaskissue.fields, "description")
                    DescriptionList.append(sDescription)
            else:
                #Create Defects
                for Defect in DefectList:
                    bCreate = True
                    for Description in DescriptionList:
                        if Description in Defect.sDescription:
                            bCreate = False;
                    if Defect.sDescription != "" and bCreate == True:
                        issue_dict = {
                            'project': {'id': 10500},
                            'summary': Defect.sDescription,
                            'description': '',
                            'issuetype': {'name': 'Defect'},
                            'parent' : { 'key' : linkissue.key},
                            'customfield_13220': int(Defect.nQty),
                            'customfield_11203' : Defect.sDescription,
                            'customfield_11029' : Defect.sDwgZone,
                            'customfield_13219' : int(sTotalQuantity), #qty recieved = total qty recieved
                            'customfield_13225' : int(sTotalQuantity), #sample size = total qty recieved
                            'customfield_11034' : Defect.sShouldBe,
                            'customfield_11035' : Defect.sActual,
                        }
                        Defectissue = jira.create_issue(fields=issue_dict)
                        DefectData = {}
                        DefectData['customfield_13220'] = int(Defect.nQty)
                        DefectData['customfield_11203'] = Defect.sDescription
                        DefectData['customfield_11029'] = Defect.sDwgZone
                        DefectData['customfield_13219'] = int(sTotalQuantity) #qty recieved = total qty recieved
                        DefectData['customfield_13225'] = int(sTotalQuantity) #sample size = total qty recieved
                        DefectData['customfield_11034'] = Defect.sShouldBe
                        DefectData['customfield_11035'] = Defect.sActual
                        DefectData['customfield_10601'] = {'value':'Use As Is'}
                        DefectData['customfield_11301'] = {'value':sCustomerApproval}
                        DefectData['customfield_12000'] = {'value':'Supplier'}
                        DefectData['customfield_12001'] = {"value" : "Mechanical","child": {"value":"300 - Dim"}}
                        jira.create_issue_link("SRMRA/MRB Link",issue,Defectissue)
                        Defectissue.update(fields=DefectData)
                        Defectissue = jira.issue(Defectissue.key)
                        if str(Defectissue.fields.status) == "Open":
                            #Create NCMR Transition
                            print("Disposition Defect")  
                            jira.transition_issue(Defectissue, transition='141')
                        #Defectissue.update(fields=data)
print("Upload Completed")  
